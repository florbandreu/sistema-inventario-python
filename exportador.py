from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from database import obtener_productos



def exportar_excel():

    productos = obtener_productos()


    if len(productos) == 0:

        print("No hay productos para exportar.")

        return


    libro = Workbook()

    hoja = libro.active

    hoja.title = "Inventario"


    encabezados = [
        "ID",
        "Nombre",
        "Categoría",
        "Precio",
        "Stock"
    ]


    hoja.append(encabezados)


    for producto in productos:

        hoja.append(producto)



    # Negrita en encabezados

    for celda in hoja[1]:

        celda.font = Font(bold=True)



    # Formato moneda

    for celda in hoja["D"][1:]:

        celda.number_format = '$#,##0.00'



    # Ajustar ancho columnas

    for columna in hoja.columns:

        largo = 0

        letra = get_column_letter(columna[0].column)


        for celda in columna:

            if celda.value:

                largo = max(largo, len(str(celda.value)))


        hoja.column_dimensions[letra].width = largo + 3



    # Mantener encabezado visible

    hoja.freeze_panes = "A2"



    libro.save("inventario.xlsx")


    print("Inventario exportado correctamente.")